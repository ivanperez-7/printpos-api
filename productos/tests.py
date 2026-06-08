from django.contrib.auth.models import User
from django.urls import reverse
from rest_framework.test import APITestCase, APIRequestFactory
from rest_framework import status

from movimiento.models import Movimiento, MovimientoItem, DetalleSalida
from system.models import RegistroActividad
from organizacion.models import Cliente, EquipoCliente, PerfilUsuario, Sucursal
from .models import Marca, Equipo, Categoría, Proveedor, Producto, Lote, Unidad
from .serializers import (
    CategoriaSerializer, MarcaSerializer, EquipoSerializer, ProveedorSerializer,
    ProductoSerializer, LoteSerializer, UnidadSerializer
)


def _create_operativo():
    user = User.objects.create_user(username='oper', password='pass')
    PerfilUsuario.objects.create(usuario=user, rol='operativo')
    return user


class MarcaModelTest(APITestCase):
    def test_str(self):
        marca = Marca.objects.create(nombre="Canon")
        self.assertEqual(str(marca), "Canon")


class EquipoModelTest(APITestCase):
    def test_str(self):
        marca = Marca.objects.create(nombre="HP")
        equipo = Equipo.objects.create(nombre="LaserJet", marca=marca)
        self.assertEqual(str(equipo), "LaserJet (HP)")


class CategoriaModelTest(APITestCase):
    def test_str(self):
        categoria = Categoría.objects.create(nombre="Impresoras")
        self.assertEqual(str(categoria), "Impresoras")


class ProveedorModelTest(APITestCase):
    def test_str(self):
        proveedor = Proveedor.objects.create(nombre="Proveedor1")
        self.assertEqual(str(proveedor), "Proveedor1")


class ProductoModelTest(APITestCase):
    def test_str(self):
        categoria = Categoría.objects.create(nombre="Cartuchos")
        proveedor = Proveedor.objects.create(nombre="Proveedor2")
        producto = Producto.objects.create(
            codigo_interno="P001",
            descripcion="Cartucho Negro",
            categoria=categoria,
            unidad_medida="pieza",
            sku="SKU001",
            min_stock=10,
            proveedor=proveedor
        )
        self.assertEqual(str(producto), "P001 (Cartucho Negro)")


class LoteModelTest(APITestCase):
    def test_str(self):
        categoria = Categoría.objects.create(nombre="Cartuchos")
        proveedor = Proveedor.objects.create(nombre="Proveedor3")
        producto = Producto.objects.create(
            codigo_interno="P002",
            descripcion="Cartucho Color",
            categoria=categoria,
            unidad_medida="pieza",
            sku="SKU002",
            min_stock=5,
            proveedor=proveedor
        )
        lote = Lote.objects.create(
            producto=producto,
            codigo_lote="L001",
            cantidad_inicial=100,
            sucursal_id=1,
        )
        self.assertEqual(str(lote), "L001")


class UnidadModelTest(APITestCase):
    def test_str(self):
        categoria = Categoría.objects.create(nombre="Cartuchos")
        proveedor = Proveedor.objects.create(nombre="Proveedor4")
        producto = Producto.objects.create(
            codigo_interno="P003",
            descripcion="Cartucho XL",
            categoria=categoria,
            unidad_medida="pieza",
            sku="SKU003",
            min_stock=3,
            proveedor=proveedor
        )
        lote = Lote.objects.create(
            producto=producto,
            codigo_lote="L002",
            cantidad_inicial=50,
            sucursal_id=1,
        )
        unidad = Unidad.objects.create(lote=lote)
        self.assertIn("Unidad de P003, lote L002:", str(unidad))


class CategoriaSerializerTest(APITestCase):
    def test_serializer(self):
        categoria = Categoría.objects.create(nombre="Tinta")
        serializer = CategoriaSerializer(categoria)
        self.assertEqual(serializer.data['nombre'], "Tinta")


class MarcaSerializerTest(APITestCase):
    def test_serializer(self):
        marca = Marca.objects.create(nombre="Epson")
        serializer = MarcaSerializer(marca)
        self.assertEqual(serializer.data['nombre'], "Epson")


class EquipoSerializerTest(APITestCase):
    def test_serializer(self):
        marca = Marca.objects.create(nombre="Brother")
        equipo = Equipo.objects.create(nombre="HL-1234", marca=marca)
        serializer = EquipoSerializer(equipo)
        self.assertEqual(serializer.data['nombre'], "HL-1234")
        self.assertEqual(serializer.data['marca']['nombre'], "Brother")


class ProveedorSerializerTest(APITestCase):
    def test_serializer(self):
        proveedor = Proveedor.objects.create(nombre="Proveedor5")
        serializer = ProveedorSerializer(proveedor)
        self.assertEqual(serializer.data['nombre'], "Proveedor5")


class ProductoSerializerTest(APITestCase):
    def test_serializer(self):
        categoria = Categoría.objects.create(nombre="Papel")
        proveedor = Proveedor.objects.create(nombre="Proveedor6")
        producto = Producto.objects.create(
            codigo_interno="P004",
            descripcion="Papel Bond",
            categoria=categoria,
            unidad_medida="caja",
            sku="SKU004",
            min_stock=20,
            proveedor=proveedor
        )
        user = _create_operativo()

        factory = APIRequestFactory()
        request = factory.post('/')
        request.user = user
        request.branch_id = 1

        serializer = ProductoSerializer(producto, context={'request': request})
        self.assertEqual(serializer.data['codigo_interno'], "P004")
        self.assertEqual(serializer.data['categoria']['nombre'], "Papel")
        self.assertEqual(serializer.data['proveedor']['nombre'], "Proveedor6")


class LoteSerializerTest(APITestCase):
    def test_serializer(self):
        categoria = Categoría.objects.create(nombre="Papel")
        proveedor = Proveedor.objects.create(nombre="Proveedor7")
        producto = Producto.objects.create(
            codigo_interno="P005",
            descripcion="Papel Reciclado",
            categoria=categoria,
            unidad_medida="paquete",
            sku="SKU005",
            min_stock=15,
            proveedor=proveedor
        )
        sucursal = Sucursal.objects.create(nombre="Suc Lote Test")
        lote = Lote.objects.create(
            producto=producto,
            codigo_lote="L003",
            cantidad_inicial=200,
            sucursal=sucursal,
        )
        user = _create_operativo()

        factory = APIRequestFactory()
        request = factory.post('/')
        request.user = user
        request.branch_id = 1
        serializer = LoteSerializer(lote, context={'request': request})
        self.assertEqual(serializer.data['codigo_lote'], "L003")


class UnidadSerializerTest(APITestCase):
    def test_serializer(self):
        categoria = Categoría.objects.create(nombre="Papel")
        proveedor = Proveedor.objects.create(nombre="Proveedor8")
        producto = Producto.objects.create(
            codigo_interno="P006",
            descripcion="Papel Fotográfico",
            categoria=categoria,
            unidad_medida="paquete",
            sku="SKU006",
            min_stock=8,
            proveedor=proveedor
        )
        lote = Lote.objects.create(
            producto=producto,
            codigo_lote="L004",
            cantidad_inicial=30,
            sucursal_id=1,
        )
        unidad = Unidad.objects.create(lote=lote)
        serializer = UnidadSerializer(unidad)
        self.assertEqual(serializer.data['lote'], lote.id)


class ProductoViewSetTest(APITestCase):
    def setUp(self):
        self.categoria = Categoría.objects.create(nombre="Accesorios")
        self.proveedor = Proveedor.objects.create(nombre="Proveedor9")
        self.producto = Producto.objects.create(
            codigo_interno="P007",
            descripcion="Cable USB",
            categoria=self.categoria,
            unidad_medida="pieza",
            sku="SKU007",
            min_stock=50,
            proveedor=self.proveedor
        )
        self.sucursal = Sucursal.objects.create(nombre="Sucursal Test")

    def test_list(self):
        user = User.objects.create_user(username='testuser', password='testpass')
        PerfilUsuario.objects.create(usuario=user, rol='admin')
        user.profile.sucursales.add(self.sucursal)
        self.client.force_login(user)

        url = reverse('producto-list')
        response = self.client.get(url, HTTP_X_BRANCH_ID=self.sucursal.id)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data[0]['codigo_interno'], "P007")


class LoteViewSetTest(APITestCase):
    def setUp(self):
        self.categoria = Categoría.objects.create(nombre="Accesorios")
        self.proveedor = Proveedor.objects.create(nombre="Proveedor10")
        self.producto = Producto.objects.create(
            codigo_interno="P008",
            descripcion="Cable HDMI",
            categoria=self.categoria,
            unidad_medida="pieza",
            sku="SKU008",
            min_stock=30,
            proveedor=self.proveedor
        )
        self.sucursal = Sucursal.objects.create(nombre="Sucursal Test 2")
        self.lote = Lote.objects.create(
            producto=self.producto,
            codigo_lote="L005",
            cantidad_inicial=60,
            sucursal=self.sucursal,
        )

    def test_list(self):
        user = User.objects.create_user(username='testuser', password='testpass')
        PerfilUsuario.objects.create(usuario=user, rol='admin')
        user.profile.sucursales.add(self.sucursal)
        self.client.force_login(user)

        url = reverse('lote-list')
        response = self.client.get(url, HTTP_X_BRANCH_ID=self.sucursal.id)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data[0]['codigo_lote'], "L005")


# ── Additional Product Model Tests ────────────────────────────────────


class ProductoModelSaveTest(APITestCase):
    def test_save_rejects_inactive_with_movements(self):
        categoria = Categoría.objects.create(nombre="Consumibles")
        proveedor = Proveedor.objects.create(nombre="Prov Save")
        producto = Producto.objects.create(
            codigo_interno="P-SAVE",
            descripcion="Test",
            categoria=categoria,
            unidad_medida="pieza",
            sku="SKU-SAVE",
            min_stock=5,
            proveedor=proveedor,
        )
        user = User.objects.create_user(username='saveuser', password='pass')
        sucursal = Sucursal.objects.create(nombre='Suc Save')
        movimiento = Movimiento.objects.create(tipo='entrada', creado_por=user, sucursal=sucursal)
        MovimientoItem.objects.create(movimiento=movimiento, producto=producto, cantidad=1)

        with self.assertRaises(ValueError):
            producto.status = 'inactivo'
            producto.save()

    def test_save_allows_inactive_without_movements(self):
        categoria = Categoría.objects.create(nombre="Consumibles2")
        proveedor = Proveedor.objects.create(nombre="Prov Save2")
        producto = Producto.objects.create(
            codigo_interno="P-SAFE",
            descripcion="Test Safe",
            categoria=categoria,
            unidad_medida="pieza",
            sku="SKU-SAFE",
            min_stock=5,
            proveedor=proveedor,
        )
        producto.status = 'inactivo'
        producto.save()
        producto.refresh_from_db()
        self.assertEqual(producto.status, 'inactivo')


# ── EquipoViewSet Custom Actions ──────────────────────────────────────


class EquipoViewSetTest(APITestCase):
    def setUp(self):
        self.user = User.objects.create_user(username='euser', password='pass')
        self.sucursal = Sucursal.objects.create(nombre='Suc Equipo')
        PerfilUsuario.objects.create(usuario=self.user, rol='admin')
        self.user.profile.sucursales.add(self.sucursal)
        self.client.force_login(self.user)
        self.marca = Marca.objects.create(nombre='Marca E')
        self.equipo = Equipo.objects.create(nombre='EQ-TEST', marca=self.marca)
        self.headers = {'HTTP_X_BRANCH_ID': self.sucursal.id}

    def test_clientes_action(self):
        cliente = Cliente.objects.create(nombre='Cli Eq', sucursal=self.sucursal)
        EquipoCliente.objects.create(
            equipo=self.equipo, cliente=cliente, alias='Alias1', contador_uso=100
        )
        url = reverse('equipo-clientes', kwargs={'pk': self.equipo.pk})
        response = self.client.get(url, **self.headers)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 1)

    def test_stats_action(self):
        cliente = Cliente.objects.create(nombre='Cli Stats', sucursal=self.sucursal)
        EquipoCliente.objects.create(
            equipo=self.equipo, cliente=cliente, alias='Stats1', contador_uso=200
        )
        url = reverse('equipo-stats', kwargs={'pk': self.equipo.pk})
        response = self.client.get(url, **self.headers)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn('total_instalaciones', response.data)
        self.assertIn('uso_total', response.data)


# ── Dashboard ─────────────────────────────────────────────────────────


class DashboardViewTest(APITestCase):
    def setUp(self):
        self.user = User.objects.create_user(username='duser', password='pass')
        self.sucursal = Sucursal.objects.create(nombre='Suc Dash')
        PerfilUsuario.objects.create(usuario=self.user, rol='admin')
        self.user.profile.sucursales.add(self.sucursal)
        self.client.force_login(self.user)
        self.headers = {'HTTP_X_BRANCH_ID': self.sucursal.id}

    def test_dashboard_returns_expected_keys(self):
        response = self.client.get('/api/v1/productos/dashboard/', **self.headers)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn('stats', response.data)
        self.assertIn('categoriasChart', response.data)
        self.assertIn('movimientosChart', response.data)
        self.assertIn('topProductosChart', response.data)
        self.assertIn('productosBajos', response.data)

    def test_dashboard_stats_include_zero_counts(self):
        response = self.client.get('/api/v1/productos/dashboard/', **self.headers)
        stats = response.data['stats']
        self.assertEqual(stats['productos'], 0)
        self.assertEqual(stats['lotes'], 0)
        self.assertEqual(stats['categorias'], 0)
        self.assertEqual(stats['proveedores'], 0)
        self.assertEqual(stats['clientes'], 0)


# ── Additional ViewSet CRUD Tests ─────────────────────────────────────


class UnidadViewSetTest(APITestCase):
    def setUp(self):
        self.user = User.objects.create_user(username='uuser', password='pass')
        self.sucursal = Sucursal.objects.create(nombre='Suc Unidad')
        PerfilUsuario.objects.create(usuario=self.user, rol='admin')
        self.user.profile.sucursales.add(self.sucursal)
        self.client.force_login(self.user)
        self.headers = {'HTTP_X_BRANCH_ID': self.sucursal.id}
        cat = Categoría.objects.create(nombre='Cat Unidad')
        prov = Proveedor.objects.create(nombre='Prov Unidad')
        self.producto = Producto.objects.create(
            codigo_interno='P-UNI', descripcion='Test', categoria=cat,
            unidad_medida='pieza', sku='SKU-UNI', min_stock=1, proveedor=prov,
        )
        self.lote = Lote.objects.create(producto=self.producto, codigo_lote='L-UNI', cantidad_inicial=5, sucursal=self.sucursal)
        self.unidad = Unidad.objects.create(lote=self.lote)

    def test_list(self):
        url = reverse('unidad-list')
        response = self.client.get(url, **self.headers)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_retrieve(self):
        url = reverse('unidad-detail', kwargs={'pk': self.unidad.pk})
        response = self.client.get(url, **self.headers)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_create(self):
        url = reverse('unidad-list')
        data = {'lote': self.lote.pk}
        response = self.client.post(url, data, format='json', **self.headers)
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)


class CategoriaViewSetTest(APITestCase):
    def setUp(self):
        self.user = User.objects.create_user(username='cuser', password='pass')
        self.sucursal = Sucursal.objects.create(nombre='Suc Cat')
        PerfilUsuario.objects.create(usuario=self.user, rol='admin')
        self.user.profile.sucursales.add(self.sucursal)
        self.client.force_login(self.user)
        self.headers = {'HTTP_X_BRANCH_ID': self.sucursal.id}
        self.categoria = Categoría.objects.create(nombre='Cat ViewSet')

    def test_list(self):
        url = reverse('categoría-list')
        response = self.client.get(url, **self.headers)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_create(self):
        url = reverse('categoría-list')
        data = {'nombre': 'Nueva Cat'}
        response = self.client.post(url, data, format='json', **self.headers)
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)


class MarcaViewSetTest(APITestCase):
    def setUp(self):
        self.user = User.objects.create_user(username='muser', password='pass')
        self.sucursal = Sucursal.objects.create(nombre='Suc Marca')
        PerfilUsuario.objects.create(usuario=self.user, rol='admin')
        self.user.profile.sucursales.add(self.sucursal)
        self.client.force_login(self.user)
        self.headers = {'HTTP_X_BRANCH_ID': self.sucursal.id}
        self.marca = Marca.objects.create(nombre='Marca VS')

    def test_list(self):
        url = reverse('marca-list')
        response = self.client.get(url, **self.headers)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_create(self):
        url = reverse('marca-list')
        data = {'nombre': 'Nueva Marca'}
        response = self.client.post(url, data, format='json', **self.headers)
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)


class ProveedorViewSetTest(APITestCase):
    def setUp(self):
        self.user = User.objects.create_user(username='puser', password='pass')
        self.sucursal = Sucursal.objects.create(nombre='Suc Prov')
        PerfilUsuario.objects.create(usuario=self.user, rol='admin')
        self.user.profile.sucursales.add(self.sucursal)
        self.client.force_login(self.user)
        self.headers = {'HTTP_X_BRANCH_ID': self.sucursal.id}
        self.proveedor = Proveedor.objects.create(nombre='Prov VS')

    def test_list(self):
        url = reverse('proveedor-list')
        response = self.client.get(url, **self.headers)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_create(self):
        url = reverse('proveedor-list')
        data = {'nombre': 'Nuevo Prov'}
        response = self.client.post(url, data, format='json', **self.headers)
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)


class CambioAnticipadoTest(APITestCase):
    def setUp(self):
        self.admin = User.objects.create_user(username='admin_ca', password='pass')
        PerfilUsuario.objects.create(usuario=self.admin, rol='admin')
        self.operativo = User.objects.create_user(username='op_ca', password='pass')
        PerfilUsuario.objects.create(usuario=self.operativo, rol='operativo')

        self.sucursal = Sucursal.objects.create(nombre='Suc CA')
        self.admin.profile.sucursales.add(self.sucursal)
        self.operativo.profile.sucursales.add(self.sucursal)

        self.categoria = Categoría.objects.create(nombre='Cat CA')
        self.proveedor = Proveedor.objects.create(nombre='Prov CA')
        self.marca = Marca.objects.create(nombre='Marca CA')
        self.equipo = Equipo.objects.create(nombre='EQ-CA', marca=self.marca)
        self.producto = Producto.objects.create(
            codigo_interno='P-CA', descripcion='Test CA',
            categoria=self.categoria, unidad_medida='pieza',
            sku='SKU-CA', min_stock=1, proveedor=self.proveedor,
            vida_util=100,
        )

        self.lote = Lote.objects.create(
            producto=self.producto, codigo_lote='L-CA',
            cantidad_inicial=5, sucursal=self.sucursal,
        )
        Unidad.objects.bulk_create([Unidad(lote=self.lote) for _ in range(5)])

        self.cliente = Cliente.objects.create(nombre='Cli CA', sucursal=self.sucursal)
        self.equipo_cliente = EquipoCliente.objects.create(
            equipo=self.equipo, cliente=self.cliente,
            alias='CA-Alias', contador_uso=50,
        )

        prior_mov = Movimiento.objects.create(
            tipo='salida', creado_por=self.operativo, sucursal=self.sucursal,
        )
        DetalleSalida.objects.create(movimiento=prior_mov, cliente=self.cliente, tecnico='Prior', subtipo='renta')
        MovimientoItem.objects.create(
            movimiento=prior_mov, producto=self.producto, cantidad=1,
            lote=self.lote, equipo_cliente=self.equipo_cliente,
            contador_uso_snapshot=0,
        )

        self.headers = {'HTTP_X_BRANCH_ID': self.sucursal.id}

    def _create_salida(self, user, extra_item_data=None):
        self.client.force_login(user)
        item_data = {
            'producto_id': self.producto.pk,
            'cantidad': 1,
            'lote_id': self.lote.pk,
            'equipo_cliente_id': self.equipo_cliente.pk,
        }
        if extra_item_data:
            item_data.update(extra_item_data)

        data = {
            'tipo': 'salida',
            'items': [item_data],
            'detalle_salida': {
                'cliente_id': self.cliente.pk,
                'tecnico': 'Test',
                'subtipo': 'renta',
            },
        }
        return self.client.post(
            reverse('movimientos-list'),
            data, format='json', **self.headers,
        )

    def test_salida_sin_flag_rechaza_aprobacion(self):
        response = self._create_salida(self.operativo)
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        mov_id = response.data['id']

        self.client.force_login(self.admin)
        url = reverse('movimientos-aprobar', args=[mov_id])
        response = self.client.post(url, **self.headers)
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_salida_con_flag_aprueba(self):
        response = self._create_salida(self.operativo, {
            'cambio_anticipado': True,
            'motivo_cambio': 'Urgencia cliente',
        })
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        mov_id = response.data['id']

        self.client.force_login(self.admin)
        url = reverse('movimientos-aprobar', args=[mov_id])
        response = self.client.post(url, **self.headers)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        registro = RegistroActividad.objects.filter(accion='approve').latest('id')
        self.assertIn('con cambios anticipados', registro.descripcion)

    def test_salida_flag_sin_motivo_rechaza(self):
        response = self._create_salida(self.operativo, {
            'cambio_anticipado': True,
            'motivo_cambio': '',
        })
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)


class RendimientoTest(APITestCase):
    def setUp(self):
        self.admin = User.objects.create_user(username='admin_rend', password='pass')
        PerfilUsuario.objects.create(usuario=self.admin, rol='admin')
        self.sucursal = Sucursal.objects.create(nombre='Suc Rend')
        self.admin.profile.sucursales.add(self.sucursal)
        self.client.force_login(self.admin)
        self.headers = {'HTTP_X_BRANCH_ID': self.sucursal.id}

        self.categoria = Categoría.objects.create(nombre='Cat Rend')
        self.proveedor = Proveedor.objects.create(nombre='Prov Rend')
        self.marca = Marca.objects.create(nombre='Marca Rend')
        self.equipo = Equipo.objects.create(nombre='EQ-Rend', marca=self.marca)
        self.producto = Producto.objects.create(
            codigo_interno='P-REND', descripcion='Toner Rend',
            categoria=self.categoria, unidad_medida='pieza',
            sku='SKU-REND', min_stock=1, proveedor=self.proveedor,
            vida_util=100,
        )
        self.lote = Lote.objects.create(
            producto=self.producto, codigo_lote='L-REND',
            cantidad_inicial=5, sucursal=self.sucursal,
        )
        self.cliente = Cliente.objects.create(nombre='Cli Rend', sucursal=self.sucursal)
        self.equipo_cliente = EquipoCliente.objects.create(
            equipo=self.equipo, cliente=self.cliente,
            alias='Rend-Alias', contador_uso=240,
        )

    def _salida(self, snapshot):
        mov = Movimiento.objects.create(
            tipo='salida', creado_por=self.admin, sucursal=self.sucursal, aprobado=True,
        )
        DetalleSalida.objects.create(movimiento=mov, cliente=self.cliente)
        MovimientoItem.objects.create(
            movimiento=mov, producto=self.producto, cantidad=1,
            lote=self.lote, equipo_cliente=self.equipo_cliente,
            contador_uso_snapshot=snapshot,
        )

    def test_rendimiento_calcula_ratio(self):
        # snapshots 0 -> 120 -> 240 => dos deltas de 120; vida_util 100 => ratio 1.2
        self._salida(0)
        self._salida(120)
        self._salida(240)

        response = self.client.get('/api/v1/productos/rendimiento/', **self.headers)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 1)

        fila = response.data[0]
        self.assertEqual(fila['producto_id'], self.producto.pk)
        self.assertEqual(fila['vida_util'], 100)
        self.assertEqual(fila['ciclos'], 2)
        self.assertEqual(fila['uso_promedio'], 120.0)
        self.assertEqual(fila['ratio'], 1.2)

    def test_rendimiento_requiere_dos_salidas(self):
        # una sola salida => no hay delta => producto ausente del reporte
        self._salida(0)
        response = self.client.get('/api/v1/productos/rendimiento/', **self.headers)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data, [])


class ExportacionTest(APITestCase):
    XLSX_CT = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    def setUp(self):
        self.admin = User.objects.create_user(username='admin_exp', password='pass')
        PerfilUsuario.objects.create(usuario=self.admin, rol='admin')
        self.sucursal = Sucursal.objects.create(nombre='Suc Exp')
        self.admin.profile.sucursales.add(self.sucursal)
        self.client.force_login(self.admin)
        self.headers = {'HTTP_X_BRANCH_ID': self.sucursal.id}

        self.categoria = Categoría.objects.create(nombre='Cat Exp')
        self.proveedor = Proveedor.objects.create(nombre='Prov Exp')
        self.producto = Producto.objects.create(
            codigo_interno='P-EXP', descripcion='Producto Exp',
            categoria=self.categoria, unidad_medida='pieza',
            sku='SKU-EXP', min_stock=2, proveedor=self.proveedor,
        )
        self.lote = Lote.objects.create(
            producto=self.producto, codigo_lote='L-EXP',
            cantidad_inicial=3, sucursal=self.sucursal,
        )
        Unidad.objects.bulk_create([Unidad(lote=self.lote) for _ in range(3)])

    def _load(self, response):
        from openpyxl import load_workbook
        import io
        return load_workbook(io.BytesIO(response.content)).active

    def test_exportar_existencias(self):
        response = self.client.get('/api/v1/productos/exportar/existencias/', **self.headers)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response['Content-Type'], self.XLSX_CT)

        ws = self._load(response)
        self.assertEqual(
            [c.value for c in ws[1]],
            ['Código', 'Descripción', 'Categoría', 'Disponible', 'Mínimo'],
        )
        fila = [c.value for c in ws[2]]
        self.assertEqual(fila[0], 'P-EXP')
        self.assertEqual(fila[3], 3)  # 3 unidades disponibles
        self.assertEqual(fila[4], 2)  # min_stock

    def test_exportar_rendimiento_vacio(self):
        response = self.client.get('/api/v1/productos/exportar/rendimiento/', **self.headers)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response['Content-Type'], self.XLSX_CT)
        ws = self._load(response)
        # Solo encabezados, sin datos de rendimiento.
        self.assertEqual(ws.max_row, 1)


class ReordenTest(APITestCase):
    def setUp(self):
        self.admin = User.objects.create_user(username='admin_reo', password='pass')
        PerfilUsuario.objects.create(usuario=self.admin, rol='admin')
        self.sucursal = Sucursal.objects.create(nombre='Suc Reo')
        self.admin.profile.sucursales.add(self.sucursal)
        self.client.force_login(self.admin)
        self.headers = {'HTTP_X_BRANCH_ID': self.sucursal.id}

        self.categoria = Categoría.objects.create(nombre='Cat Reo')
        self.proveedor = Proveedor.objects.create(nombre='Prov Reo')

    def _producto(self, codigo, min_stock, disponibles):
        producto = Producto.objects.create(
            codigo_interno=codigo, descripcion=f'Desc {codigo}',
            categoria=self.categoria, unidad_medida='pieza',
            sku=f'SKU-{codigo}', min_stock=min_stock, proveedor=self.proveedor,
        )
        lote = Lote.objects.create(
            producto=producto, codigo_lote=f'L-{codigo}',
            cantidad_inicial=disponibles, sucursal=self.sucursal,
        )
        Unidad.objects.bulk_create([Unidad(lote=lote) for _ in range(disponibles)])
        return producto

    def _consumo(self, producto, cantidad):
        # Salida aprobada solo para alimentar la estadística de consumo.
        mov = Movimiento.objects.create(
            tipo='salida', creado_por=self.admin, sucursal=self.sucursal, aprobado=True,
        )
        MovimientoItem.objects.create(movimiento=mov, producto=producto, cantidad=cantidad)

    def test_producto_bajo_minimo_aparece_con_sugerencia(self):
        # min_stock=10, disponibles=3 => bajo mínimo. Consumo 12 en 6 meses => 2/mes.
        producto = self._producto('P-REO', min_stock=10, disponibles=3)
        self._consumo(producto, 12)

        response = self.client.get('/api/v1/productos/reorden/', **self.headers)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 1)

        grupo = response.data[0]
        self.assertEqual(grupo['proveedor_nombre'], 'Prov Reo')
        self.assertEqual(len(grupo['productos']), 1)

        fila = grupo['productos'][0]
        self.assertEqual(fila['codigo_interno'], 'P-REO')
        self.assertEqual(fila['consumo_mensual'], 2.0)
        # cantidad_sugerida = round(2*2 - 3) = 1 (meses_objetivo por defecto = 2)
        self.assertEqual(fila['cantidad_sugerida'], 1)

    def test_producto_bien_abastecido_ausente(self):
        # min_stock=1, disponibles=5, sin consumo => no se sugiere reorden.
        self._producto('P-OK', min_stock=1, disponibles=5)
        response = self.client.get('/api/v1/productos/reorden/', **self.headers)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data, [])
