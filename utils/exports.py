import io

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font


def build_xlsx(headers, rows, sheet_title='Reporte'):
    """Genera un archivo .xlsx en memoria y devuelve los bytes.

    headers: lista de encabezados de columna.
    rows: iterable de filas; cada fila es una lista/tupla alineada con headers.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]  # límite de Excel para nombre de hoja

    ws.append(list(headers))
    for celda in ws[1]:
        celda.font = Font(bold=True)

    for row in rows:
        ws.append(list(row))

    # Ancho de columna aproximado al contenido más largo.
    for idx, header in enumerate(headers, start=1):
        largo_max = len(str(header))
        for row in rows:
            valor = row[idx - 1] if idx - 1 < len(row) else ''
            largo_max = max(largo_max, len(str(valor)))
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = min(largo_max + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def xlsx_response(headers, rows, filename, sheet_title='Reporte'):
    """Atajo: construye el .xlsx y lo envuelve en una HttpResponse descargable."""
    contenido = build_xlsx(headers, rows, sheet_title=sheet_title)
    return HttpResponse(
        contenido,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )
